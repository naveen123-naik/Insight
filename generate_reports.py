import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF
import os

# Load dataset
csv_path = r'c:\Users\naikn\OneDrive\Desktop\Insight\docs\sales.csv'
df = pd.read_csv(csv_path)

# Calculate key metrics
df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce').fillna(1)
df['Price'] = pd.to_numeric(df['Price'], errors='coerce').fillna(0)
df['Profit'] = pd.to_numeric(df['Profit'], errors='coerce').fillna(df['Price'] * df['Quantity'] * 0.2)
df['Total_Amount'] = df['Price'] * df['Quantity']

total_revenue = float(df['Total_Amount'].sum())
total_profit = float(df['Profit'].sum())
total_units = int(df['Quantity'].sum())
total_orders = len(df)
avg_order_val = total_revenue / total_orders if total_orders else 0

prod_summary = df.groupby('Product')['Total_Amount'].sum().reset_index().sort_values(by='Total_Amount', ascending=False)
best_seller = str(prod_summary.iloc[0]['Product']) if not prod_summary.empty else 'N/A'

city_summary = df.groupby('City')['Total_Amount'].sum().reset_index().sort_values(by='Total_Amount', ascending=False)
top_city = str(city_summary.iloc[0]['City']) if not city_summary.empty else 'N/A'

cat_summary = df.groupby('Category')['Total_Amount'].sum().reset_index().sort_values(by='Total_Amount', ascending=False)
top_category = str(cat_summary.iloc[0]['Category']) if not cat_summary.empty else 'N/A'

# -------------------------------------------------------------
# 1. GENERATE EXCEL (.xlsx) REPORT
# -------------------------------------------------------------
excel_output_path = r'c:\Users\naikn\OneDrive\Desktop\Insight\InsightAI_Executive_Report_sales.xlsx'
wb = openpyxl.Workbook()

# Sheet 1: Executive KPI Summary
ws1 = wb.active
ws1.title = "Executive KPI Summary"
ws1.views.sheetView[0].showGridLines = True

header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
card_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
border_thin = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                     top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))

# Title banner
ws1.merge_cells('A1:E2')
title_cell = ws1['A1']
title_cell.value = "InsightAI Executive Intelligence Report - sales.csv"
title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
title_cell.fill = header_fill
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# KPI Table
kpis = [
    ("Total Revenue (INR)", f"₹{total_revenue:,.2f}"),
    ("Total Net Profit (INR)", f"₹{total_profit:,.2f}"),
    ("Total Volume Sold", f"{total_units:,} Units"),
    ("Total Orders Count", f"{total_orders:,} Orders"),
    ("Average Order Value", f"₹{avg_order_val:,.2f}"),
    ("Best Selling Product", best_seller),
    ("Top Revenue Region", top_city),
    ("Dominant Category", top_category)
]

ws1.cell(row=4, column=1, value="Key Performance Indicator").font = Font(name="Calibri", size=11, bold=True, color="1E3A8A")
ws1.cell(row=4, column=2, value="Dataset Value").font = Font(name="Calibri", size=11, bold=True, color="1E3A8A")

for idx, (k, v) in enumerate(kpis, start=5):
    c1 = ws1.cell(row=idx, column=1, value=k)
    c2 = ws1.cell(row=idx, column=2, value=v)
    c1.font = Font(name="Calibri", size=11, bold=True)
    c2.font = Font(name="Calibri", size=11, color="2563EB" if "₹" in v else "0F172A", bold=True)
    c1.border = border_thin
    c2.border = border_thin
    c1.fill = card_fill

ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 30

# Sheet 2: Breakdown Analytics
ws2 = wb.create_sheet(title="Category & City Breakdown")
ws2.views.sheetView[0].showGridLines = True

ws2.cell(row=1, column=1, value="Category Revenue Breakdown").font = Font(name="Calibri", size=12, bold=True, color="1E3A8A")
ws2.cell(row=2, column=1, value="Category").font = Font(name="Calibri", size=10, bold=True)
ws2.cell(row=2, column=2, value="Revenue (INR)").font = Font(name="Calibri", size=10, bold=True)

for i, r in cat_summary.iterrows():
    row_num = 3 + i
    ws2.cell(row=row_num, column=1, value=r['Category'])
    ws2.cell(row=row_num, column=2, value=r['Total_Amount']).number_format = '"₹ "#,##0.00'

ws2.cell(row=1, column=4, value="Regional City Sales").font = Font(name="Calibri", size=12, bold=True, color="1E3A8A")
ws2.cell(row=2, column=4, value="City").font = Font(name="Calibri", size=10, bold=True)
ws2.cell(row=2, column=5, value="Revenue (INR)").font = Font(name="Calibri", size=10, bold=True)

for i, r in city_summary.iterrows():
    row_num = 3 + i
    ws2.cell(row=row_num, column=4, value=r['City'])
    ws2.cell(row=row_num, column=5, value=r['Total_Amount']).number_format = '"₹ "#,##0.00'

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['D'].width = 25
ws2.column_dimensions['E'].width = 25

# Sheet 3: Cleaned Dataset Records
ws3 = wb.create_sheet(title="Clean Dataset Records")
ws3.views.sheetView[0].showGridLines = True

headers = list(df.columns)
for col_num, h_text in enumerate(headers, 1):
    c = ws3.cell(row=1, column=col_num, value=h_text)
    c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

for r_idx, row in df.iterrows():
    for c_idx, val in enumerate(row, 1):
        ws3.cell(row=r_idx+2, column=c_idx, value=val)

wb.save(excel_output_path)
print(f"Excel Report created successfully at: {excel_output_path}")

# -------------------------------------------------------------
# 2. GENERATE PDF REPORT (.pdf)
# -------------------------------------------------------------
pdf_output_path = r'c:\Users\naikn\OneDrive\Desktop\Insight\InsightAI_Executive_Report_sales.pdf'

class PDFReport(FPDF):
    def header(self):
        self.set_fill_color(30, 58, 138)
        self.rect(0, 0, 210, 25, 'F')
        self.set_font("Helvetica", "B", 14)
        self.set_text_color(255, 255, 255)
        self.set_y(5)
        self.cell(0, 8, "InsightAI - Executive Intelligence Report", align="C", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 9)
        self.cell(0, 4, "Dataset: sales.csv (34 Cleaned Transactions)", align="C", new_x="LMARGIN", new_y="NEXT")
        self.ln(8)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(148, 163, 184)
        self.cell(0, 10, f"Page {self.page_no()} | Generated by InsightAI Platform", align="C")

pdf = PDFReport()
pdf.add_page()
pdf.set_auto_page_break(auto=True, margin=15)

# Executive Summary Box
pdf.set_fill_color(248, 250, 252)
pdf.set_draw_color(37, 99, 235)
pdf.rect(10, 30, 190, 28, 'DF')

pdf.set_xy(14, 33)
pdf.set_font("Helvetica", "B", 11)
pdf.set_text_color(15, 23, 42)
pdf.cell(0, 6, "AI Insights Executive Summary", new_x="LMARGIN", new_y="NEXT")

pdf.set_font("Helvetica", "", 9)
pdf.set_text_color(71, 85, 105)
summary_txt = f"Analyzed {total_orders} dataset records. Generated total revenue of INR {total_revenue:,.2f} and net profit of INR {total_profit:,.2f}. The top-performing product is '{best_seller}', while '{top_city}' led regional revenue performance."
pdf.set_x(14)
pdf.multi_cell(180, 4.5, summary_txt)

# KPI Cards Section
pdf.set_y(63)
pdf.set_font("Helvetica", "B", 12)
pdf.set_text_color(15, 23, 42)
pdf.cell(0, 8, "Key Financial Metrics", new_x="LMARGIN", new_y="NEXT")

kpi_data = [
    ("TOTAL REVENUE", f"INR {total_revenue:,.0f}"),
    ("TOTAL PROFIT", f"INR {total_profit:,.0f}"),
    ("VOLUME SOLD", f"{total_units:,} Units"),
    ("BEST SELLER", best_seller)
]

col_w = 44
gap = 4
start_x = 10
y_kpi = pdf.get_y()

for i, (label, val) in enumerate(kpi_data):
    x_pos = start_x + i * (col_w + gap)
    pdf.set_fill_color(241, 245, 249)
    pdf.set_draw_color(203, 213, 225)
    pdf.rect(x_pos, y_kpi, col_w, 20, 'DF')
    
    pdf.set_xy(x_pos + 2, y_kpi + 3)
    pdf.set_font("Helvetica", "B", 7)
    pdf.set_text_color(71, 85, 105)
    pdf.cell(col_w - 4, 4, label, align="L")
    
    pdf.set_xy(x_pos + 2, y_kpi + 9)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_text_color(37, 99, 235)
    pdf.cell(col_w - 4, 6, val, align="L")

pdf.set_y(y_kpi + 26)

# Strategic Takeaways
pdf.set_font("Helvetica", "B", 11)
pdf.set_text_color(15, 23, 42)
pdf.cell(0, 7, "Strategic Recommendations", new_x="LMARGIN", new_y="NEXT")

recs = [
    f"Inventory Allocation: Increase stock buffers for top-selling '{best_seller}' ahead of peak quarter.",
    f"Regional Focus: Allocate 20% additional marketing budget in '{top_city}' to capture regional demand.",
    "Margin Optimization: Review supplier cost structures for low-margin accessories to raise gross margin."
]

pdf.set_font("Helvetica", "", 9)
pdf.set_text_color(51, 65, 85)
for r in recs:
    pdf.set_x(14)
    pdf.multi_cell(180, 5, f"- {r}")

# Category Breakdown Table
pdf.ln(4)
pdf.set_font("Helvetica", "B", 11)
pdf.set_text_color(15, 23, 42)
pdf.cell(0, 7, "Category Revenue Breakdown", new_x="LMARGIN", new_y="NEXT")

pdf.set_fill_color(30, 58, 138)
pdf.set_font("Helvetica", "B", 9)
pdf.set_text_color(255, 255, 255)
pdf.cell(60, 6, "Category", border=1, fill=True)
pdf.cell(60, 6, "Revenue (INR)", border=1, fill=True)
pdf.cell(60, 6, "Contribution %", border=1, fill=True, new_x="LMARGIN", new_y="NEXT")

pdf.set_font("Helvetica", "", 9)
pdf.set_text_color(15, 23, 42)
for _, r in cat_summary.iterrows():
    share = (r['Total_Amount'] / total_revenue) * 100 if total_revenue else 0
    pdf.cell(60, 6, str(r['Category']), border=1)
    pdf.cell(60, 6, f"INR {r['Total_Amount']:,.2f}", border=1)
    pdf.cell(60, 6, f"{share:.1f}%", border=1, new_x="LMARGIN", new_y="NEXT")

pdf.output(pdf_output_path)
print(f"PDF Report created successfully at: {pdf_output_path}")
